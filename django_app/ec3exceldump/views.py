import csv
import datetime
import logging
import os
import random
import re
import threading
import time
import decimal

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse, Http404
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from ec3exceldump.models import Report, Account, Meter, RowIssue, RESOURCE_TYPE_CHOICES, STM
from .forms import ReportForm

from openpyxl import load_workbook

# Create your views here.

# Get an instance of a logger
logger = logging.getLogger(__name__)

class Echo:
	def write(self, value):
		return value

def index(request):
	reports = Report.objects.all()
	if request.method == 'POST':
		form = ReportForm(request.POST, request.FILES)
		if form.is_valid():
			cover_date = form.cleaned_data.get('cover_date', False)
			account_file = form.cleaned_data.get('account_file', False)
			meter_file = form.cleaned_data.get('meter_file', False)
			if not cover_date:
				raise forms.ValidationError('Invalid date')
			if not account_file:
				raise forms.ValidationError('Invalid account file')
			if not meter_file:
				raise forms.ValidationError('Invalid meter file')

			report = Report(cover_date=cover_date, account_file=account_file, meter_file=meter_file)
			report.save()

			t = threading.Thread(target=loadbooks, args=[report.account_file, report.meter_file, report.id])
			t.setDaemon(True)
			t.start()
			
			return redirect('/ec3')
	else:
		form = ReportForm()
	return render(request, 'index.html', { 'reports': reports, 'form': form })

@csrf_exempt
def reportupdate(request):
	if request.method == 'GET':
		raise Http404()
	result = {}
	arr = request.POST.getlist('reports[]', False);
	if not arr:
		raise Http404()

	for entry in arr:
		id = int(entry)
		obj = {}
		report = Report.objects.get(pk=id)
		if report:
			if report.account_file_crawled:
				obj['afc'] = True
				if report.account_crawl_result == 1:
					obj['acr'] = 1
				else:
					obj['acr'] = report.hex_account_crawl()
					obj['acm'] = report.account_crawl_msg
			else:
				obj['afp'] = str(report.account_file_progress / 100)
				obj['afc'] = False

			if report.meter_file_crawled:
				obj['mfc'] = True
				if report.meter_crawl_result == 1:
					obj['mcr'] = 1
				else:
					obj['mcr'] = report.hex_meter_crawl()
					obj['mcm'] = report.meter_crawl_msg
			else:
				obj['mfp'] = str(report.meter_file_progress / 100)
				obj['mfc'] = False

			if report.account_crawl_result == 1 and report.meter_crawl_result == 1:
				obj['link'] = "<a href=\"ec3/viewreport/" + str(report.id) + "\">View Report</a>"
		result['report' + entry] = obj
	return JsonResponse(result)

def viewreport(request, id):
	report = Report.objects.get(pk=id)
	accounts = Account.objects.filter(report=report, included=True)
	page = request.GET.get('page', 1)

	try:
		page = int(page)
	except ValueError:
		page = 1
	
	accounts, search = performSearch(accounts, request.GET)
	accounts = accounts.order_by('pk')
	size = accounts.count()

	p = Paginator(accounts, 100)
	page = max(1, min(page, p.num_pages))
	onPage = p.page(page)
	accounts = onPage.object_list
	if search:
		search.append("uID=" + str(random.randint(10000, 99999)))
	search = '&'.join(search)
	return render(request, 'viewreport.html', { 'report': report, 'accounts': accounts, 'size': size, 'num_pages': p.num_pages, 'page': page, 'p': onPage, 'search': search })

def exceptions(request, id):
	report = Report.objects.get(pk=id)
	if not report:
		return Http404()
	issues = RowIssue.objects.filter(report=report)
	accts = issues.filter(meter=None).order_by('excel_row')
	mtrs = issues.filter(account=None).order_by('excel_row')
	return render(request, 'exceptions.html', { 'report': report, 'accts': accts, 'mtrs': mtrs })

@csrf_exempt
def updatexception(request, id):
	issue = RowIssue.objects.get(pk=id)
	if not issue:
		raise Http404()

	val = request.POST.get('val', False)
	result = { 'success': False }
	if not val:
		result['msg'] = "Missing POST value"
	elif not isinstance(val, str):
		result['msg'] = "Missing 'val' value from POST"
	elif not val.lower()[0] in [ 'i', 'e' ]:
		result['msg'] = "'val' can only be \"Included\" or \"Excluded\""
	else:
		if val.lower()[0] == 'i':
			val = True
		else:
			val = False

		if issue.meter:
			issue.meter.included = val
			issue.meter.save()
		if issue.account:
			issue.account.included = val
			issue.account.save()
		
		result['success'] = True
		result['val'] = val

	return JsonResponse(result)

@csrf_exempt
def retrieveaccount(request, id):
	account = Account.objects.get(pk=id)
	if not account:
		raise Http404()

	result = { 'success': True }
	obj = {}
	obj['college'] = account.college
	obj['billing_period'] = account.billing_period
	obj['account_number'] = str(account.account_number)
	obj['resource_type'] = account.get_resource_type()
	obj['energy_usage'] = account.energy_usage
	obj['demand_usage'] = account.demand_usage
	obj['billed_amount'] = account.billed_amount
	result['obj'] = obj
	return JsonResponse(result)

@csrf_exempt
def retrievemeter(request, id):
	meter = Meter.objects.get(pk=id)
	if not meter:
		raise Http404()

	result = { 'success': True }
	obj = {}
	obj['meter_number'] = meter.meter_number
	obj['energy_usage'] = meter.energy_usage
	obj['demand_usage'] = meter.demand_usage
	result['obj'] = obj
	return JsonResponse(result)

def performSearch(accounts, obj):
	start = obj.get('start', False)
	end = obj.get('end', False)
	search = []
	if start and end:
		try:
			datetime.datetime.strptime(start, '%Y-%m-%d')
			datetime.datetime.strptime(end, '%Y-%m-%d')

			accounts = accounts.filter(billing_period__range=[start, end])
			search.append("start=" + str(start))
			search.append("end=" + str(end))
		except ValueError:
			pass
	return accounts, search

def createcsv(request, id):
	accounts = Account.objects.filter(report=id, included=True)
	accounts = performSearch(accounts, request.GET)[0]

	accounts = accounts.order_by('billing_period')
	echo_buffer = Echo()
	writer = csv.writer(echo_buffer)

	def rows():
		yield writer.writerow(['Campus', 'Date', 'Account Number', 'Meter Number', 'Type', 'Usage', 'Demand', 'Bill Amount'])
		for account in accounts:
			college = account.college
			billing_period = account.billing_period
			account_number = account.account_number
			resource_type = account.get_resource_type()
			yield writer.writerow([
				college,
				billing_period,
				str(account_number),
				'',
				resource_type,
				account.energy_usage,
				account.demand_usage,
				account.billed_amount])
			meters = account.meter_set.filter(included=True)
			for meter in meters:
				yield writer.writerow([
					college,
					billing_period,
					str(account_number),
					meter.meter_number,
					resource_type,
					meter.energy_usage,
					meter.demand_usage,
					''])

	response = StreamingHttpResponse(rows(), content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename="dumpity dump.csv"'
	
	return response

def loadbooks(account_path, meter_path, report_id):
	report = Report.objects.get(pk=report_id)

	logger.warning('Loading Account and Meter Books')
	account_book = load_workbook(filename=account_path, read_only=True)
	ws = account_book.active
	total = 0
	id_map = {}
	curr = time.time()
	for row in ws.iter_rows(min_row=2, values_only=True):
		total = total + 1
		included = None

		college = row[1]
		account_number = row[7]
		billing_period = row[2]
		for fmt in settings.DATE_INPUT_FORMATS:
			try:
		 		billing_period = datetime.datetime.strptime(str(billing_period), fmt)
			except ValueError:
				continue
			break

		billed_amount = row[13]
		try:
			billed_amount = round(float(billed_amount), 2)
		except ValueError:
			pass

		resource_type = row[8]
		if isinstance(resource_type, str) and len(resource_type) > 0:
			resource_type = resource_type[0]
		energy_usage = row[11]
		if energy_usage == "NULL":
			energy_usage = 0
		demand_usage = row[12]
		if demand_usage == "NULL":
			demand_usage = 0

		account = Account(
			report=report,
			college=college,
			account_number=account_number,
			billing_period=billing_period,
			billed_amount=billed_amount,
			resource_type=resource_type,
			energy_usage=energy_usage,
			demand_usage=demand_usage)
		try:
			# This will first validate all fields
			# before attempting to save. If a
			# validation fails, we can catch it
			# and create an issue.
			account.full_clean()
			account.save()
		except ValidationError as e:
			msg = []
			for k, vs in e.message_dict.items():
				if k == '__all__':
					key = ''
				else:
					setattr(account, k, None)
					key = []
					for i in k.split('_'):
						key.append(i.upper()[0] + i.lower()[1:])
					key = '(' + ' '.join(key) + ') '
				for v in vs:
					msg.append(key + v)
			msg = "\n".join(msg)
			# logger.warning(msg)
			account.included = False
			account.save()
			exp = RowIssue(report=report, account=account, excel_row=total + 1, message=msg)
			exp.save()

		if not row[2] in id_map:
			id_map[row[2]] = {}
		id_map[row[2]][row[32]] = account.id

		report.account_file_progress = int(total * 10000 / (ws.max_row - 1));
		t = time.time()
		if t - curr > 2:
			curr = t
			report.save()
	logger.warning("Saved " + str(total) + " accounts")

	report.account_file_crawled = True
	report.account_crawl_result = 1
	report.account_file_progress = 10000
	report.save()
	
	meter_book = load_workbook(filename=meter_path, read_only=True)
	exps = []
	ws = meter_book.active
	total = 0
	curr = time.time()
	for row in ws.iter_rows(min_row=2, values_only=True):
		total = total + 1
		meter_number = row[2]
		if meter_number == 'TRNSM':
			meter_number = '0'
		account_id = None
		if row[3] in id_map and row[1] in id_map[row[3]]:
			account_id = id_map[row[3]][row[1]]
		energy_usage = row[6]
		if energy_usage == 'NULL':
			energy_usage = 0
		demand_usage = row[7]
		if demand_usage == 'NULL':
			demand_usage = 0

		meter = Meter(
			report=report,
			account_id=account_id,
			meter_number=meter_number,
			energy_usage=energy_usage,
			demand_usage=demand_usage)
		try:
			# This will first validate all fields
			# before attempting to save. If a
			# validation fails, we can catch it
			# and create an issue.
			meter.full_clean()
			meter.save()
		except ValidationError as e:
			msg = []
			for k, vs in e.message_dict.items():
				if k == '__all__':
					key = ''
				else:
					setattr(meter, k, None)
					key = []
					for i in k.split('_'):
						key.append(i.upper()[0] + i.lower()[1:])
					key = '(' + ' '.join(key) + ') '
				for v in vs:
					msg.append(key + v)
			msg = "\n".join(msg)
			# logger.warning(msg)
			meter.included = False
			meter.save()
			exp = RowIssue(report=report, meter=meter, excel_row=total + 1, message=msg)
			exp.save()

		report.meter_file_progress = int(total * 10000 / (ws.max_row - 1));
		t = time.time()
		if t - curr > 2:
			curr = t
			report.save()
	logger.warning("Saved " + str(total) + " meters")

	report.meter_file_crawled = True
	report.meter_crawl_result = 1
	report.meter_file_progress = 10000
	report.save()

	logger.warning('Saved Report')